"""
Tests for TrialViewSet.export — 试验导出 xlsx endpoint (P2 阶段)

契约:
  - URL: GET /api/events/trials/export/?format=xlsx&status=...&...
  - 复用 filterset_fields(状态、设备、负责人、起止日期)
  - 权限: IsAdminOrManagerOrReadOnly(GET 允许所有已认证用户)
  - 输出: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
  - Content-Disposition: attachment; filename="trials-YYYY-MM-DD.xlsx"
  - 列: 试验名称 / 状态(中文)/ 主开始时间 / 主结束时间

注意: openpyxl 把 datetime 单元格解析为 datetime 对象(非 str),
      把空单元格解析为 None(非 '')。本测试按这一行为断言。
"""
from datetime import datetime, timedelta
from io import BytesIO

import pytest
from django.utils import timezone
from openpyxl import load_workbook

from events.models import Trial

URL = "/api/events/trials/export/"


@pytest.mark.django_db
class TestTrialExportEndpoint:
    def test_export_unauthenticated_returns_401(self, api_client):
        response = api_client.get(URL)
        assert response.status_code == 401

    def test_export_regular_user_returns_xlsx(self, api_client, regular_user_obj):
        """IsAdminOrManagerOrReadOnly 对所有已认证用户允许 GET。"""
        Trial.objects.create(
            title="Trial Regular",
            client="Client A",
            description="desc",
            status="planned",
        )
        api_client.force_authenticate(user=regular_user_obj)

        response = api_client.get(URL)
        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment" in response["Content-Disposition"]
        assert response["Content-Disposition"].endswith('.xlsx"')

    def test_export_admin_sees_all_columns(self, admin_client):
        """admin 导出一行 Trial,验证四列内容。openpyxl 把日期列解析为 datetime。

        注意:Trial.save() 会从 time_slots 重算 start_date/end_date。
        直接 .update() 绕过 save,显式控制时间字段。
        """
        now = timezone.now()
        trial = Trial.objects.create(
            title="Trial Admin",
            client="Client",
            description="desc",
            status="in_progress",
        )
        Trial.objects.filter(pk=trial.pk).update(
            start_date=now + timedelta(days=1),
            end_date=now + timedelta(days=3),
        )

        response = admin_client.get(URL)
        assert response.status_code == 200

        wb = load_workbook(filename=BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # 表头 + 1 行数据
        assert rows[0] == ("试验名称", "状态", "主开始时间", "主结束时间")
        assert len(rows) == 2
        title_col, status_col, start_col, end_col = rows[1]
        assert title_col == "Trial Admin"
        assert status_col == "进行中"  # 中文 label,不是 "in_progress"
        assert isinstance(start_col, datetime)
        assert isinstance(end_col, datetime)

    def test_export_applies_status_filter(self, admin_client):
        """?status=completed 应只返回状态为 completed 的 trial。"""
        Trial.objects.create(title="Planned Trial", client="C", description="d", status="planned")
        Trial.objects.create(title="Completed Trial", client="C", description="d", status="completed")
        Trial.objects.create(title="Cancelled Trial", client="C", description="d", status="cancelled")

        response = admin_client.get(URL + "?status=completed")
        assert response.status_code == 200

        wb = load_workbook(filename=BytesIO(response.content))
        ws = wb.active
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))

        assert len(data_rows) == 1
        assert data_rows[0][0] == "Completed Trial"
        assert data_rows[0][1] == "已完成"

    def test_export_applies_start_date_filter(self, admin_client):
        """?start_date__gte= 应应用 filterset_fields 中的过滤,只返回匹配 trial。

        Trial.save() 会从 time_slots 重算 start_date,所以用 .update() 绕过 save。
        """
        cutoff = timezone.now() + timedelta(days=1)
        future = Trial.objects.create(title="Future Trial", client="C", description="d", status="planned")
        past = Trial.objects.create(title="Past Trial", client="C", description="d", status="planned")
        Trial.objects.filter(pk=future.pk).update(
            start_date=cutoff + timedelta(days=10),
            end_date=cutoff + timedelta(days=12),
        )
        Trial.objects.filter(pk=past.pk).update(
            start_date=cutoff - timedelta(days=10),
            end_date=cutoff - timedelta(days=8),
        )

        # 仅返回 start_date >= cutoff 的(Future Trial)
        # 注意 cutoff.isoformat() 包含 `+00:00`,URL 中 `+` 必须编码为 %2B,
        # 否则 DRF 测试客户端会把 `+` 解码为空格,过滤条件解析失败。
        from urllib.parse import quote
        cutoff_iso = cutoff.isoformat()
        cutoff_qs = quote(cutoff_iso, safe="")
        response = admin_client.get(URL + f"?start_date__gte={cutoff_qs}")
        assert response.status_code == 200

        wb = load_workbook(filename=BytesIO(response.content))
        ws = wb.active
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))

        titles = [row[0] for row in data_rows]
        assert titles == ["Future Trial"]

    def test_export_empty_database_returns_header_only(self, admin_client):
        """空库导出只有表头,不算错误。"""
        response = admin_client.get(URL)
        assert response.status_code == 200

        wb = load_workbook(filename=BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 1
        assert rows[0] == ("试验名称", "状态", "主开始时间", "主结束时间")

    def test_export_null_dates_render_as_empty_cell(self, admin_client):
        """start_date / end_date 为 null 时,Excel 单元格为空(不是字符串 'None')。"""
        Trial.objects.create(
            title="No Date Trial",
            client="C",
            description="d",
            status="planned",
            start_date=None,
            end_date=None,
        )

        response = admin_client.get(URL)
        assert response.status_code == 200

        wb = load_workbook(filename=BytesIO(response.content))
        ws = wb.active
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 1
        # openpyxl 把空单元格解析为 None(不是字符串 ''也不是 'None')
        assert data_rows[0][2] is None
        assert data_rows[0][3] is None

    def test_export_filename_includes_today_date(self, admin_client):
        """Content-Disposition 文件名含当天日期。"""
        Trial.objects.create(title="Trial X", client="C", description="d", status="planned")

        response = admin_client.get(URL)
        today = timezone.now().date().isoformat()
        assert f'trials-{today}.xlsx' in response["Content-Disposition"]