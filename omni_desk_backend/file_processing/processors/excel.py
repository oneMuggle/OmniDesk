import openpyxl
from typing import Dict, Any
from .base import FileProcessor


class ExcelProcessor(FileProcessor):
    """Excel 文件处理器"""

    def extract_text(self, file_path: str) -> str:
        """提取 Excel 文本（所有 Sheet 拼接）"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        text_parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_parts.append(f"# {sheet_name}\n")

            for row in ws.iter_rows(values_only=True):
                row_text = '\t'.join(str(cell) if cell is not None else '' for cell in row)
                text_parts.append(row_text)
            text_parts.append('\n')

        return '\n'.join(text_parts)

    def extract_markdown(self, file_path: str) -> str:
        """提取 Markdown 表格"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        md_parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            md_parts.append(f"## {sheet_name}\n\n")

            data = list(ws.iter_rows(values_only=True))
            if not data:
                continue

            headers = data[0]
            md_parts.append('| ' + ' | '.join(str(h) if h else '' for h in headers) + ' |\n')
            md_parts.append('| ' + ' | '.join('---' for _ in headers) + ' |\n')

            for row in data[1:]:
                md_parts.append('| ' + ' | '.join(str(cell) if cell else '' for cell in row) + ' |\n')

            md_parts.append('\n')

        return ''.join(md_parts)

    def extract_structured(self, file_path: str) -> Dict[str, Any]:
        """提取结构化数据（多 Sheet）"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            data = list(ws.iter_rows(values_only=True))

            if not data:
                continue

            headers = data[0]
            rows = data[1:]

            sheets.append({
                'name': sheet_name,
                'headers': [str(h) if h else f'列{i+1}' for i, h in enumerate(headers)],
                'data': [[str(cell) if cell is not None else '' for cell in row] for row in rows],
                'row_count': len(rows),
                'column_count': len(headers),
            })

        return {
            'sheet_count': len(sheets),
            'sheets': sheets,
        }

    def get_metadata(self, file_path: str) -> Dict[str, Any]:
        """获取 Excel 元数据"""
        wb = openpyxl.load_workbook(file_path, read_only=True)
        return {
            'sheet_count': len(wb.sheetnames),
            'sheet_names': wb.sheetnames,
        }
