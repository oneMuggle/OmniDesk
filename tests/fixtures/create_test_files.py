import os
from openpyxl import Workbook

os.makedirs('tests/fixtures', exist_ok=True)

# 创建简单 Excel
wb = Workbook()
ws = wb.active
ws.title = "Sheet1"
ws.append(['姓名', '年龄', '城市'])
ws.append(['张三', 25, '北京'])
ws.append(['李四', 30, '上海'])
ws.append(['王五', 28, '北京'])
wb.save('tests/fixtures/simple.xlsx')

# 创建多 Sheet Excel
wb2 = Workbook()
ws1 = wb2.active
ws1.title = "销售数据"
ws1.append(['月份', '销售额', '成本'])
ws1.append(['1月', 10000, 6000])
ws1.append(['2月', 15000, 8000])

ws2 = wb2.create_sheet("员工信息")
ws2.append(['姓名', '部门', '薪资'])
ws2.append(['张三', '销售', 8000])
wb2.save('tests/fixtures/multi_sheet.xlsx')

print("Test files created successfully")
